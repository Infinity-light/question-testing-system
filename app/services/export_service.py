import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from flask import current_app
from app.models import TestResult, Question


class ExportService:
    """Service for exporting test results to Excel"""

    def export_to_excel(self, test_result_ids: list, output_filename: str = None) -> str:
        """
        Export test results to Excel file matching the template format.

        Args:
            test_result_ids: List of test result IDs to export
            output_filename: Optional custom filename

        Returns:
            Path to the generated Excel file
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Results"

        # Define headers (10 columns as per template)
        headers = [
            "标题",           # 1. Title
            "题目类型",       # 2. Question type
            "领域",           # 3. Subject
            "难度",           # 4. Difficulty
            "知识点",         # 5. Knowledge points
            "问题",           # 6. Question text
            "答案",           # 7. Standard answer
            "解题思路",       # 8. Solution approach
            "",               # 9. Empty column
            "查难情况"        # 10. Difficulty status (X/8)
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Set column widths
        ws.column_dimensions['A'].width = 20  # 标题
        ws.column_dimensions['B'].width = 15  # 题目类型
        ws.column_dimensions['C'].width = 15  # 领域
        ws.column_dimensions['D'].width = 10  # 难度
        ws.column_dimensions['E'].width = 30  # 知识点
        ws.column_dimensions['F'].width = 50  # 问题
        ws.column_dimensions['G'].width = 30  # 答案
        ws.column_dimensions['H'].width = 50  # 解题思路
        ws.column_dimensions['I'].width = 5   # Empty
        ws.column_dimensions['J'].width = 15  # 查难情况

        # Populate data rows
        row_num = 2
        for test_result_id in test_result_ids:
            test_result = TestResult.query.get(test_result_id)
            if not test_result:
                current_app.logger.warning(f"Test result {test_result_id} not found")
                continue

            question = test_result.question

            # Write data
            ws.cell(row=row_num, column=1, value=question.title)
            ws.cell(row=row_num, column=2, value=question.question_type)
            ws.cell(row=row_num, column=3, value=question.subject)
            ws.cell(row=row_num, column=4, value=question.difficulty)
            ws.cell(row=row_num, column=5, value=question.knowledge_points)
            ws.cell(row=row_num, column=6, value=question.question_text)
            ws.cell(row=row_num, column=7, value=question.standard_answer)
            ws.cell(row=row_num, column=8, value=question.solution_approach)
            ws.cell(row=row_num, column=9, value="")  # Empty column
            ws.cell(row=row_num, column=10, value=test_result.difficulty_status)

            # Apply text wrapping for long content
            for col in [5, 6, 7, 8]:
                ws.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True, vertical='top')

            row_num += 1

        # Generate filename
        if not output_filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"test_results_{timestamp}.xlsx"

        # Ensure export directory exists
        export_dir = current_app.config['EXPORT_DIR']
        os.makedirs(export_dir, exist_ok=True)

        # Save file
        output_path = os.path.join(export_dir, output_filename)
        wb.save(output_path)

        current_app.logger.info(f"Exported {len(test_result_ids)} test results to {output_path}")

        return output_path

    def export_qualified_questions(self, output_filename: str = None) -> str:
        """
        Export only qualified questions (success rate < 50%) to Excel.

        Args:
            output_filename: Optional custom filename

        Returns:
            Path to the generated Excel file
        """
        # Get all qualified test results
        qualified_results = TestResult.query.filter_by(qualified=True).all()
        test_result_ids = [result.id for result in qualified_results]

        if not test_result_ids:
            current_app.logger.warning("No qualified test results found")
            return None

        return self.export_to_excel(test_result_ids, output_filename)


# Global service instance
export_service = ExportService()
